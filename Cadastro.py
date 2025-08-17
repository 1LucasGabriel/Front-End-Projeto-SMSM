from flask import Flask, render_template, request, redirect
from openpyxl import Workbook, load_workbook
import os

ARQUIVO = "funcionarios.xlsx"
app = Flask(__name__)

# if planilha exist == false -> planilha exist == true
if not os.path.exists(ARQUIVO):
    wb = Workbook()
    for sheet in ["Médicos", "Enfermeiros", "Outros"]:
        ws = wb.create_sheet(sheet)
        ws.append(["Nome Completo", "Celular", "Email Institucional", "CRM", "COREN"])
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(ARQUIVO)


@app.route("/", methods=["GET", "POST"])
def cadastro():
    if request.method == "POST":
        cargo = request.form["cargo"]
        nome = request.form["nome"]
        celular = request.form["celular"]
        email = request.form["email"]
        crm = request.form.get("crm", "")
        coren = request.form.get("coren", "")

        # abre planilha e escolhe aba
        wb = load_workbook(ARQUIVO)
        if cargo == "Médico":
            ws = wb["Médicos"]
        elif cargo == "Enfermeiro":
            ws = wb["Enfermeiros"]
        else:
            ws = wb["Outros"]

        ws.append([nome, celular, email, crm, coren])
        wb.save(ARQUIVO)

        return redirect("/")

    return render_template("form.html")


@app.route("/lista")
def lista():
    wb = load_workbook(ARQUIVO)
    dados = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        funcionarios = list(ws.iter_rows(min_row=2, values_only=True))
        dados[sheet] = funcionarios

    return render_template("lista.html", dados=dados)


if __name__ == "__main__":
    app.run(debug=True)
